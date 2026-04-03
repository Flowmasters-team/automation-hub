"""
Заполнение отчёта РАО об использованных произведениях (форма ВГТРК).

Генерирует Excel-таблицу (.xlsx) с точной структурой формы РАО:
  Приложение №1 к Лицензионному договору между РАО и ВГТРК.

Колонки:
  1. Наименование передачи
  2. Дата и время выхода в эфир (число, часы, мин.)
  3. Название музыкальных произведений, используемых в программе
  4. ФИО композитора
  5. ФИО автора текста
  6. Длительность звучания произведения (мин. сек.)
  7. Количество исполнений
  8. Общий хронометраж
  9. Жанр произведения
  10. Исполнитель (ФИО исполнителя или название коллектива)

Зависимости: pip install mutagen openpyxl
"""

import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, numbers
except ImportError:
    print("Ошибка: не установлена библиотека openpyxl")
    print("Установите: pip install openpyxl")
    sys.exit(1)

from extract_metadata import extract_from_list, extract_from_directory


# --- Колонки формы РАО ---
RAO_COLUMNS = [
    ("Наименование передачи", 30),
    ("Дата и время выхода в эфир\n(число, часы, мин.)", 18),
    ("Название музыкальных произведений,\nиспользуемых в программе", 35),
    ("ФИО\nкомпозитора", 20),
    ("ФИО автора\nтекста", 18),
    ("Длительность\nзвучания\nпроизведения\n(мин. сек.)", 14),
    ("Количест-\nво\nисполне-\nний", 12),
    ("Общий\nхрономет-\nраж", 12),
    ("Жанр\nпроизведе-\nния", 14),
    ("Исполнитель\n(ФИО исполнителя\nили название\nколлектива)", 25),
]


def _fmt_duration(seconds: float) -> str:
    """Форматирует длит��льность как M:SS:ss или H:MM:SS."""
    total = int(seconds)
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    if h > 0:
        return f"{h}:{m:02d}:{s:02d}"
    return f"0:{m:02d}:{s:02d}"


def _seconds_to_excel_time(seconds: float) -> float:
    """Конвертирует секунды в дробное число Excel (доля суток)."""
    return seconds / 86400.0


def _multiply_duration(seconds: float, count: int) -> str:
    """Общий хронометраж = длительност�� * кол-во исполнений."""
    return _fmt_duration(seconds * count)


def create_rao_report(
    programs: list[dict],
    output_path: str,
    month: str = "",
    quarter: str = "",
    year: str = "",
    contract_no: str = "",
    contractor: str = "",
):
    """
    Создаёт Excel-файл в формате отчёта РАО.

    programs — список программ, каждая:
    {
        "name": "АМВ Вездино Церковь ч1",
        "air_date": "04.09.25",
        "tracks": [
            {
                "title": "Gothic Oblivion",     # из метаданных или вручную
                "composer": "PP Music (UK)",     # ФИО композитора
                "lyricist": "",                  # ФИО автора текста
                "duration_seconds": 76,          # из метаданных
                "play_count": 1,                 # кол-во исполнений (по умолчанию 1)
                "genre": "пьеса",                # жанр
                "performer": "инстр. Анс",       # исполнитель
            },
            ...
        ]
    }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    # --- Стили ---
    font_main = Font(name="Times New Roman", size=10)
    font_bold = Font(name="Times New Roman", size=10, bold=True)
    font_title = Font(name="Times New Roman", size=12, bold=True)
    font_header = Font(name="Times New Roman", size=9, bold=True)
    font_rules = Font(name="Times New Roman", size=9)
    font_side = Font(name="Times New Roman", size=10, bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
    align_vertical = Alignment(text_rotation=90, horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_left = Border(left=thin)
    border_right = Border(right=thin)
    border_bottom = Border(bottom=thin)
    border_lr = Border(left=thin, right=thin)

    col_offset = 3  # данные начинаются с колонки D (index 4)
    MIN_EMPTY_ROWS = 30  # минимум пустых строк после данных (как в оригинале)

    # --- Ширины колонок ---
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 4
    for i, (_, w) in enumerate(RAO_COLUMNS):
        col_letter = chr(ord("D") + i)
        ws.column_dimensions[col_letter].width = w

    # ============================================================
    # ШАПКА: Приложение №1
    # ============================================================
    row = 2
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=13)
    ws.cell(row=row, column=10, value="Приложение № 1").font = font_main
    row = 3
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=13)
    ws.cell(row=row, column=10, value="к Лицензионному договору").font = font_main
    row = 4
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=13)
    contract_text = f'от «__» ____20__г. № {contract_no}/ТВ' if contract_no else 'от «__» ____20__г. № ____/ТВ'
    ws.cell(row=row, column=10, value=contract_text).font = font_main
    row = 5
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=13)
    contractor_text = f"между РАО и {contractor}" if contractor else "между РАО и ________________________"
    ws.cell(row=row, column=10, value=contractor_text).font = font_main
    row = 6
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=13)
    ws.cell(row=row, column=10, value="_______________________________________").font = font_main

    # ============================================================
    # ЗАГОЛОВОК: ОТЧЕТ
    # ============================================================
    row = 8
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=13)
    ws.cell(row=row, column=4, value="ОТЧЕТ").font = font_title
    ws.cell(row=row, column=4).alignment = align_center

    row = 9
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=13)
    ws.cell(row=row, column=4, value="об использованных  произведениях").font = font_bold
    ws.cell(row=row, column=4).alignment = align_center

    row = 10
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=13)
    if not month:
        month = datetime.now().strftime("%B")
    if not year:
        year = datetime.now().strftime("%Y")
    period_text = f"за _____{month}_____(месяц) квартал  {year}  г."
    ws.cell(row=row, column=4, value=period_text).font = font_main
    ws.cell(row=row, column=4).alignment = align_center

    # ============================================================
    # ЗАГОЛОВКИ ТАБЛИЦЫ (строки 11-12) + номера (13)
    # ============================================================
    header_row_start = 11
    for col_i, (col_name, _) in enumerate(RAO_COLUMNS):
        col_idx = col_offset + 1 + col_i
        ws.merge_cells(
            start_row=header_row_start, start_column=col_idx,
            end_row=header_row_start + 1, end_column=col_idx,
        )
        cell = ws.cell(row=header_row_start, column=col_idx, value=col_name)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_all
        ws.cell(row=header_row_start + 1, column=col_idx).border = border_all

    num_row = header_row_start + 2  # строка 13
    for col_i in range(len(RAO_COLUMNS)):
        col_idx = col_offset + 1 + col_i
        cell = ws.cell(row=num_row, column=col_idx, value=col_i + 1)
        cell.font = font_main
        cell.alignment = align_center
        cell.border = border_all

    # ============================================================
    # ДАННЫЕ
    # ============================================================
    data_row = num_row + 1  # строка 14

    for prog in programs:
        prog_name = prog.get("name", "")
        air_date = prog.get("air_date", "")
        tracks = prog.get("tracks", [])

        for t_idx, track in enumerate(tracks):
            title = track.get("title", "")
            composer = track.get("composer", "")
            lyricist = track.get("lyricist", "")
            dur_sec = track.get("duration_seconds", 0)
            play_count = track.get("play_count", 1)
            genre = track.get("genre", "пьеса")
            performer = track.get("performer", "")

            dur_excel = _seconds_to_excel_time(dur_sec)
            total_excel = _seconds_to_excel_time(dur_sec * play_count)
            time_fmt = 'h:mm:ss'

            row_values = [
                prog_name if t_idx == 0 else "",
                air_date if t_idx == 0 else "",
                title, composer, lyricist,
                dur_excel, play_count, total_excel,
                genre, performer,
            ]

            for col_i, val in enumerate(row_values):
                col_idx = col_offset + 1 + col_i
                cell = ws.cell(row=data_row, column=col_idx, value=val)
                cell.font = font_main
                cell.alignment = align_top_center if col_i in (1, 5, 6, 7) else align_top_left
                cell.border = border_all
                if col_i in (5, 7):
                    cell.number_format = time_fmt

            data_row += 1

    # ============================================================
    # ПУСТЫЕ СТРОКИ С РАМКОЙ (как в оригинале — таблица продолжается)
    # ============================================================
    last_data_row = data_row
    for empty_r in range(last_data_row, last_data_row + MIN_EMPTY_ROWS):
        for col_i in range(len(RAO_COLUMNS)):
            col_idx = col_offset + 1 + col_i
            ws.cell(row=empty_r, column=col_idx).border = border_all
    data_row = last_data_row + MIN_EMPTY_ROWS + 2

    # ============================================================
    # ПРАВИЛА ЗАПОЛНЕНИЯ ОТЧЕТА (с рамкой вокруг всего блока)
    # ============================================================
    rules_start = data_row
    rules = [
        "* Приложение № 1 заполняется для произведений, используемых в передачах, анонсах, в качестве музыкального оформления, заставках и т.п.;",
        "* данные по анонсам предоставляются отдельным отчетом по форме Приложения № 1;",
        "* отчет составляется на бумажном носителе (рукописное заполнение не допускается) и в электронном виде;",
        "* форма отчета не подлежит корректировке, дополнение и удаление граф не допускается; все графы формы обязательны к заполнению, в случае если в передаче используется фрагмент фильма без музыки, необходимо это указать;",
        "* информация в отчете указывается в строгом соответствии с наименованием граф;",
        "* отчет сортируется в алфавитном порядке по названию передачи, а внутри передачи - по музыкальным произведениям;",
        "* отчет заполняется на русском языке с использованием кириллицы, использование латинских букв при написании российских произведений и авторов не допускается;",
        "* для произведений авторов-субъектов РФ, при указании данных на языке оригинала (национальном языке) в скобках в обязательном порядке указывается перевод на русский язык;",
        "* для произведений иностранных авторов данные указываются на языке оригинала и без сокращений;",
        "* отчет составляется с использованием шрифта размером не менее 12 пунктов;",
        "* все страницы отчета должны быть пронумерованы и заверены печатью и подписью.",
    ]
    rules_end = rules_start + len(rules)  # last rule row (inclusive)

    # Header row: top + left + right borders (merged D:M)
    ws.merge_cells(start_row=rules_start, start_column=4, end_row=rules_start, end_column=13)
    cell_rh = ws.cell(row=rules_start, column=4, value="Правила заполнения отчета:")
    cell_rh.font = font_bold
    cell_rh.border = Border(left=thin, right=thin, top=thin)

    # Rule rows: left + right borders on merged cell (top-left of range);
    # last row also gets bottom border
    for i, rule in enumerate(rules):
        r = rules_start + 1 + i
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=13)
        cell = ws.cell(row=r, column=4, value=rule)
        cell.font = font_rules
        cell.alignment = align_top_left
        if r == rules_end:
            cell.border = Border(left=thin, right=thin, bottom=thin)
        else:
            cell.border = border_lr

    # ============================================================
    # БОКОВЫЕ ПОДПИСИ: ОБЩЕСТВО (вверху, как в оригинале)
    # ============================================================
    # Колонка A: "ОБЩЕСТВО" вертикально (строки 2-6)
    ws.merge_cells(start_row=2, start_column=1, end_row=6, end_column=1)
    cell_o = ws.cell(row=2, column=1, value="ОБЩЕСТВО")
    cell_o.font = font_side
    cell_o.alignment = align_vertical

    # Колонка B: "ОБЩЕСТВО" вертикально (строки 2-6)
    ws.merge_cells(start_row=2, start_column=2, end_row=6, end_column=2)
    cell_o2 = ws.cell(row=2, column=2, value="ОБЩЕСТВО")
    cell_o2.font = font_side
    cell_o2.alignment = align_vertical

    # Увеличиваем высоту строк 2-6 чтобы вертикальный текст поместился
    for r in range(2, 7):
        ws.row_dimensions[r].height = 20

    # М.П. (строка 7)
    ws.cell(row=7, column=1, value="М.П.").font = font_main

    # Строка 8: М.П. в колонке B (как в оригинале — рядом с ОТЧЕТ)
    ws.cell(row=8, column=2, value="М.П.").font = font_main

    # Колонка C: "(И.А.Базилевский)" вертикально (строки 9-13, рядом с заголовками таблицы)
    ws.merge_cells(start_row=9, start_column=3, end_row=13, end_column=3)
    cell_baz = ws.cell(row=9, column=3, value="(И.А.Базилевский)")
    cell_baz.font = font_main
    cell_baz.alignment = align_vertical

    # ============================================================
    # БОКОВЫЕ ПОДПИСИ: ПОЛЬЗОВАТЕЛЬ (внизу, как в оригинале)
    # ============================================================
    # ПОЛЬЗОВАТЕЛЬ вертикально — 8 строк перед правилами
    footer_side_start = rules_start - 6
    ws.merge_cells(start_row=footer_side_start, start_column=1, end_row=footer_side_start + 8, end_column=1)
    cell_p = ws.cell(row=footer_side_start, column=1, value="ПОЛЬЗОВАТЕЛЬ")
    cell_p.font = font_side
    cell_p.alignment = align_vertical

    # БАЗИЛЕВСКИЙ в колонке B вертикально (рядом с ПОЛЬЗОВАТЕЛЬ)
    ws.merge_cells(start_row=footer_side_start, start_column=2, end_row=footer_side_start + 8, end_column=2)
    cell_baz2 = ws.cell(row=footer_side_start, column=2, value="БАЗИЛЕВСКИЙ")
    cell_baz2.font = font_side
    cell_baz2.alignment = align_vertical

    # М.П. (под ПОЛЬЗОВАТЕЛЬ)
    mp_row = footer_side_start + 9
    ws.cell(row=mp_row, column=1, value="М.П.").font = font_main

    # (                           ) в колонке B
    ws.cell(row=mp_row + 1, column=1, value="(                           )").font = font_main

    # ============================================================
    # ПОДПИСИ ВНИЗУ
    # ============================================================
    sign_row = rules_end + 4

    ws.merge_cells(start_row=sign_row, start_column=4, end_row=sign_row, end_column=6)
    cell_mp = ws.cell(row=sign_row, column=4, value="          М.П. _________________")
    cell_mp.font = font_main
    cell_mp.alignment = align_left

    ws.merge_cells(start_row=sign_row, start_column=7, end_row=sign_row, end_column=8)
    ws.cell(row=sign_row, column=7, value="_____________________").font = font_main

    ws.merge_cells(start_row=sign_row, start_column=10, end_row=sign_row, end_column=11)
    ws.cell(row=sign_row, column=10, value="Дата_____________").font = font_main

    # Подписи под линиями
    ws.merge_cells(start_row=sign_row + 1, start_column=4, end_row=sign_row + 1, end_column=6)
    cell_sig = ws.cell(row=sign_row + 1, column=4, value="               (подпись)")
    cell_sig.font = font_main
    cell_sig.alignment = align_left

    ws.merge_cells(start_row=sign_row + 1, start_column=7, end_row=sign_row + 1, end_column=8)
    ws.cell(row=sign_row + 1, column=7, value="(должность, ФИО руководителя)").font = font_main

    # --- Сохранение ---
    wb.save(output_path)
    return output_path


def tracks_to_program(
    tracks: list[dict],
    program_name: str,
    air_date: str,
    default_genre: str = "пьеса",
    default_performer: str = "инстр. Анс",
) -> dict:
    """Преобразует список метаданных треков в формат программы для отчёта."""
    program_tracks = []
    for t in tracks:
        composer = t.get("artist", "") or t.get("composer", "")
        program_tracks.append({
            "title": t.get("title", ""),
            "composer": composer,
            "lyricist": "",
            "duration_seconds": t.get("duration_seconds", 0),
            "play_count": 1,
            "genre": t.get("genre", "") or default_genre,
            "performer": t.get("artist", "") or default_performer,
        })
    return {
        "name": program_name,
        "air_date": air_date,
        "tracks": program_tracks,
    }


# --- CLI ---
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Заполнение отчёта РАО об использованных произведениях (ВГТРК)"
    )
    parser.add_argument(
        "files", nargs="*",
        help="Пути к аудиофайлам"
    )
    parser.add_argument(
        "--dir", "-d",
        help="Директория с аудиофайлами"
    )
    parser.add_argument(
        "--list", "-l",
        help="Текстовый файл со списком путей (по одному на строку)"
    )
    parser.add_argument(
        "--output", "-o", default="rao_report.xlsx",
        help="Имя выходного файла (по умолчанию: rao_report.xlsx)"
    )
    parser.add_argument(
        "--program", "-p", default="",
        help="Название программы/передачи"
    )
    parser.add_argument(
        "--date", default="",
        help="Дата эфира (например: 04.09.25)"
    )
    parser.add_argument(
        "--month", default="",
        help="Месяц отчёта (например: сентябрь)"
    )
    parser.add_argument(
        "--year", default="",
        help="Год отчёта (например: 2025)"
    )
    parser.add_argument(
        "--genre", default="пьеса",
        help="Жанр по умолчанию (по умолчанию: пьеса)"
    )
    parser.add_argument(
        "--performer", default="инстр. Анс",
        help="Исполнитель по умолчанию (по умолчанию: инстр. Анс)"
    )
    parser.add_argument(
        "--json", dest="json_input",
        help="JSON-файл с полной структурой (несколько программ)"
    )

    args = parser.parse_args()

    # --- Режим 1: JSON с несколькими программами ---
    if args.json_input:
        with open(args.json_input, encoding="utf-8") as f:
            data = json.load(f)
        programs = data.get("programs", data if isinstance(data, list) else [data])
        output = create_rao_report(
            programs,
            args.output,
            month=data.get("month", args.month),
            quarter=data.get("quarter", ""),
            year=data.get("year", args.year),
        )
        print(f"Отчёт РАО сохранён: {os.path.abspath(output)}")
        sys.exit(0)

    # --- Режим 2: аудиофайлы ---
    file_paths = []
    if args.files:
        file_paths.extend(args.files)
    if args.list:
        with open(args.list, encoding="utf-8") as f:
            file_paths.extend(line.strip() for line in f if line.strip())

    if args.dir:
        tracks = extract_from_directory(args.dir)
    elif file_paths:
        tracks = extract_from_list(file_paths)
    else:
        parser.print_help()
        print("\nОшибка: укажите файлы, --dir, --list или --json")
        sys.exit(1)

    if not tracks:
        print("Не найдено аудиофайлов.")
        sys.exit(1)

    # Выводим что нашли
    print(f"\nНайдено треков: {len(tracks)}")
    for i, t in enumerate(tracks, 1):
        err = t.get("error", "")
        status = f"!! {err}" if err else "OK"
        artist = t.get("artist") or t.get("composer") or "?"
        print(f"  {i}. {artist} — {t.get('title', '?')} [{t.get('duration', '?')}] ({status})")

    # Формируем программу
    program = tracks_to_program(
        tracks,
        program_name=args.program or "Без названия",
        air_date=args.date or datetime.now().strftime("%d.%m.%y"),
        default_genre=args.genre,
        default_performer=args.performer,
    )

    output = create_rao_report(
        [program],
        args.output,
        month=args.month,
        year=args.year,
    )
    print(f"\nОтчёт РАО сохранён: {os.path.abspath(output)}")
    print("\nПодсказка: для нескольких передач используйте --json с файлом структуры.")
    print("Пример JSON: python fill_form.py --json example_report.json")
